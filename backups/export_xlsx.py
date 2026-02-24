"""
Genera un archivo Excel (.xlsx) con todos los datos del negocio.
Hojas: Ventas, Detalle Ventas, Productos, Inventario, Mov. Stock, Gastos, Caja, Empleados.
"""
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# -- Style constants --
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
MONEY_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD/MM/YYYY'
DATETIME_FORMAT = 'DD/MM/YYYY HH:MM'
THIN_BORDER = Border(
    bottom=Side(style='thin', color='E5E7EB'),
)


def _style_header(ws, col_count):
    """Apply styles to the header row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def _auto_width(ws, min_width=10, max_width=40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


def _write_sheet(ws, headers, rows, money_cols=None, date_cols=None, datetime_cols=None):
    """Write headers + data rows with formatting."""
    money_cols = money_cols or []
    date_cols = date_cols or []
    datetime_cols = datetime_cols or []

    # Headers
    ws.append(headers)
    _style_header(ws, len(headers))

    # Data
    for row_data in rows:
        ws.append(row_data)

    # Format columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
            col_idx = cell.column
            if col_idx in money_cols:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal='right')
            elif col_idx in date_cols:
                cell.number_format = DATE_FORMAT
            elif col_idx in datetime_cols:
                cell.number_format = DATETIME_FORMAT

    _auto_width(ws)
    ws.freeze_panes = 'A2'


def generate_export(tenant, days=None):
    """
    Generate an Excel workbook with all business data.
    Returns an openpyxl Workbook.
    If days is set, only exports data from the last N days.
    """
    from accounting.models import CashMovement, CashRegister, Expense
    from employees.models import Employee
    from inventory.models import Ingredient, StockMovement
    from products.models import Category, Product, ProductVariant
    from sales.models import Sale, SaleItem

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    since = None
    if days:
        since = timezone.now() - timedelta(days=days)

    # =========================================================
    # Sheet 1: Ventas
    # =========================================================
    ws = wb.create_sheet('Ventas')
    headers = [
        'Numero', 'Fecha', 'Cliente', 'Tipo', 'Estado',
        'Subtotal', 'IVA', 'Descuento', 'Delivery', 'Total',
        'Metodo Pago', 'Referencia', 'Pagado', 'Vendedor',
    ]

    sales_qs = Sale.objects.filter(tenant=tenant).select_related(
        'payment_method', 'created_by'
    ).order_by('-created_at')
    if since:
        sales_qs = sales_qs.filter(created_at__gte=since)

    tipo_map = {'local': 'Local', 'takeaway': 'Para Llevar', 'delivery': 'Delivery'}
    status_map = {
        'pending': 'Pendiente', 'preparing': 'Preparando', 'ready': 'Listo',
        'delivered': 'Entregado', 'cancelled': 'Cancelado',
    }

    rows = []
    for s in sales_qs:
        rows.append([
            s.sale_number,
            s.created_at.replace(tzinfo=None) if s.created_at else '',
            s.customer_name or '',
            tipo_map.get(s.order_type, s.order_type),
            status_map.get(s.status, s.status),
            float(s.subtotal),
            float(s.tax_amount),
            float(s.discount_amount),
            float(s.delivery_fee),
            float(s.total_amount),
            s.payment_method.name if s.payment_method else '',
            s.payment_reference or '',
            'Si' if s.is_paid else 'No',
            s.created_by.get_full_name() or s.created_by.username if s.created_by else '',
        ])
    _write_sheet(ws, headers, rows,
                 money_cols=[6, 7, 8, 9, 10],
                 datetime_cols=[2])

    # =========================================================
    # Sheet 2: Detalle Ventas
    # =========================================================
    ws = wb.create_sheet('Detalle Ventas')
    headers = [
        'Nro Venta', 'Fecha', 'Producto', 'Variantes', 'Cantidad',
        'Precio Unit', 'Total Linea', 'Notas',
    ]

    items_qs = SaleItem.objects.filter(sale__tenant=tenant).select_related(
        'sale', 'product'
    ).order_by('-sale__created_at')
    if since:
        items_qs = items_qs.filter(sale__created_at__gte=since)

    # Pre-fetch variant names
    all_variant_ids = set()
    items_list = list(items_qs)
    for item in items_list:
        if item.selected_variants:
            for vid in item.selected_variants:
                try:
                    all_variant_ids.add(int(vid))
                except (ValueError, TypeError):
                    pass
    variant_names = {}
    if all_variant_ids:
        for v in ProductVariant.objects.filter(id__in=all_variant_ids):
            variant_names[v.id] = v.name

    rows = []
    for item in items_list:
        variants_str = ''
        if item.selected_variants:
            names = [variant_names.get(int(vid), f'#{vid}')
                     for vid in item.selected_variants
                     if str(vid).isdigit()]
            variants_str = ', '.join(names)

        rows.append([
            item.sale.sale_number,
            item.sale.created_at.replace(tzinfo=None) if item.sale.created_at else '',
            item.product.name if item.product else '',
            variants_str,
            item.quantity,
            float(item.unit_price),
            float(item.unit_price * item.quantity),
            item.notes or '',
        ])
    _write_sheet(ws, headers, rows,
                 money_cols=[6, 7],
                 datetime_cols=[2])

    # =========================================================
    # Sheet 3: Productos
    # =========================================================
    ws = wb.create_sheet('Productos')
    headers = [
        'Nombre', 'Categoria', 'Precio', 'Tiene Variantes',
        'Requiere Preparacion', 'Activo', 'Destacado',
    ]

    products = Product.objects.filter(tenant=tenant).select_related('category').order_by('category__sort_order', 'name')

    rows = []
    for p in products:
        rows.append([
            p.name,
            p.category.name if p.category else '',
            float(p.base_price),
            'Si' if p.has_variants else 'No',
            'Si' if p.requires_preparation else 'No',
            'Si' if p.is_active else 'No',
            'Si' if p.is_featured else 'No',
        ])
    _write_sheet(ws, headers, rows, money_cols=[3])

    # Sheet 3b: Variantes (sub-sheet)
    ws = wb.create_sheet('Variantes')
    headers = ['Producto', 'Tipo', 'Nombre', 'Modificador Precio', 'Por Defecto', 'Activo']

    variants = ProductVariant.objects.filter(
        product__tenant=tenant
    ).select_related('product').order_by('product__name', 'variant_type', 'sort_order')

    tipo_variant_map = {
        'size': 'Tamano', 'flavor': 'Sabor', 'topping': 'Extra/Topping',
        'preparation': 'Preparacion', 'custom': 'Personalizado',
    }
    rows = []
    for v in variants:
        rows.append([
            v.product.name,
            tipo_variant_map.get(v.variant_type, v.variant_type),
            v.name,
            float(v.price_modifier),
            'Si' if v.is_default else 'No',
            'Si' if v.is_active else 'No',
        ])
    _write_sheet(ws, headers, rows, money_cols=[4])

    # =========================================================
    # Sheet 4: Inventario
    # =========================================================
    ws = wb.create_sheet('Inventario')
    headers = [
        'Ingrediente', 'Unidad', 'Stock Actual', 'Stock Minimo',
        'Costo Unitario', 'Valor Stock', 'Proveedor', 'Estado',
    ]

    ingredients = Ingredient.objects.filter(tenant=tenant).select_related('supplier').order_by('name')

    rows = []
    for ing in ingredients:
        rows.append([
            ing.name,
            ing.get_unit_display(),
            float(ing.current_stock),
            float(ing.min_stock),
            float(ing.cost_per_unit),
            float(ing.stock_value),
            ing.supplier.name if ing.supplier else '',
            'BAJO' if ing.is_low_stock else 'OK',
        ])
    _write_sheet(ws, headers, rows, money_cols=[5, 6])

    # =========================================================
    # Sheet 5: Movimientos de Stock
    # =========================================================
    ws = wb.create_sheet('Mov. Stock')
    headers = [
        'Fecha', 'Ingrediente', 'Tipo', 'Cantidad', 'Costo Unit',
        'Costo Total', 'Notas', 'Usuario',
    ]

    mov_type_map = {
        'purchase': 'Compra', 'usage': 'Uso', 'adjustment': 'Ajuste',
        'waste': 'Desperdicio', 'return': 'Devolucion',
    }

    movements_qs = StockMovement.objects.filter(
        ingredient__tenant=tenant
    ).select_related('ingredient', 'created_by').order_by('-created_at')
    if since:
        movements_qs = movements_qs.filter(created_at__gte=since)

    rows = []
    for m in movements_qs:
        rows.append([
            m.created_at.replace(tzinfo=None) if m.created_at else '',
            m.ingredient.name if m.ingredient else '',
            mov_type_map.get(m.movement_type, m.movement_type),
            float(m.quantity),
            float(m.unit_cost) if m.unit_cost else 0,
            float(m.total_cost) if m.total_cost else 0,
            m.notes or '',
            m.created_by.get_full_name() or m.created_by.username if m.created_by else '',
        ])
    _write_sheet(ws, headers, rows,
                 money_cols=[5, 6],
                 datetime_cols=[1])

    # =========================================================
    # Sheet 6: Gastos
    # =========================================================
    ws = wb.create_sheet('Gastos')
    headers = [
        'Fecha', 'Categoria', 'Descripcion', 'Monto',
        'Pagado por', 'Nro Recibo', 'Notas',
    ]

    expenses_qs = Expense.objects.filter(tenant=tenant).select_related(
        'category', 'paid_by'
    ).order_by('-date')
    if since:
        expenses_qs = expenses_qs.filter(date__gte=since.date())

    rows = []
    for e in expenses_qs:
        rows.append([
            e.date,
            e.category.name if e.category else '',
            e.description,
            float(e.amount),
            e.paid_by.get_full_name() or e.paid_by.username if e.paid_by else '',
            e.receipt_number or '',
            e.notes or '',
        ])
    _write_sheet(ws, headers, rows,
                 money_cols=[4],
                 date_cols=[1])

    # =========================================================
    # Sheet 7: Caja
    # =========================================================
    ws = wb.create_sheet('Caja')
    headers = [
        'Fecha', 'Apertura', 'Cierre', 'Esperado', 'Diferencia',
        'Estado', 'Abrio', 'Cerro', 'Notas',
    ]

    registers_qs = CashRegister.objects.filter(tenant=tenant).select_related(
        'opened_by', 'closed_by'
    ).order_by('-date')
    if since:
        registers_qs = registers_qs.filter(date__gte=since.date())

    status_caja_map = {'open': 'Abierta', 'closed': 'Cerrada'}

    rows = []
    for r in registers_qs:
        rows.append([
            r.date,
            float(r.opening_amount),
            float(r.closing_amount) if r.closing_amount is not None else '',
            float(r.expected_amount) if r.expected_amount is not None else '',
            float(r.difference) if r.difference is not None else '',
            status_caja_map.get(r.status, r.status),
            r.opened_by.get_full_name() or r.opened_by.username if r.opened_by else '',
            r.closed_by.get_full_name() or r.closed_by.username if r.closed_by else '',
            r.notes or '',
        ])
    _write_sheet(ws, headers, rows,
                 money_cols=[2, 3, 4, 5],
                 date_cols=[1])

    # =========================================================
    # Sheet 8: Movimientos de Caja
    # =========================================================
    ws = wb.create_sheet('Mov. Caja')
    headers = [
        'Fecha', 'Tipo', 'Monto', 'Descripcion', 'Referencia', 'Usuario',
    ]

    cash_type_map = {
        'sale': 'Venta', 'expense': 'Gasto', 'withdrawal': 'Retiro',
        'deposit': 'Deposito', 'adjustment': 'Ajuste', 'tip': 'Propina',
    }

    cash_movs_qs = CashMovement.objects.filter(
        register__tenant=tenant
    ).select_related('register', 'created_by').order_by('-created_at')
    if since:
        cash_movs_qs = cash_movs_qs.filter(created_at__gte=since)

    rows = []
    for cm in cash_movs_qs:
        rows.append([
            cm.created_at.replace(tzinfo=None) if cm.created_at else '',
            cash_type_map.get(cm.movement_type, cm.movement_type),
            float(cm.amount),
            cm.description or '',
            cm.reference or '',
            cm.created_by.get_full_name() or cm.created_by.username if cm.created_by else '',
        ])
    _write_sheet(ws, headers, rows,
                 money_cols=[3],
                 datetime_cols=[1])

    # =========================================================
    # Sheet 9: Empleados
    # =========================================================
    ws = wb.create_sheet('Empleados')
    headers = [
        'Nombre', 'Apellido', 'Puesto', 'Telefono',
        'Email', 'DNI', 'Salario Mensual', 'Desde', 'Activo',
    ]

    position_map = {
        'encargado': 'Encargado', 'cajero': 'Cajero', 'cocinero': 'Cocinero',
        'pizzero': 'Pizzero', 'delivery': 'Delivery', 'mozo': 'Mozo',
        'limpieza': 'Limpieza', 'ayudante': 'Ayudante',
    }

    employees = Employee.objects.filter(tenant=tenant).order_by('last_name', 'first_name')

    rows = []
    for emp in employees:
        rows.append([
            emp.first_name,
            emp.last_name,
            position_map.get(emp.position, emp.position),
            emp.phone or '',
            emp.email or '',
            emp.dni or '',
            float(emp.monthly_salary) if emp.monthly_salary else 0,
            emp.hire_date,
            'Si' if emp.is_active else 'No',
        ])
    _write_sheet(ws, headers, rows,
                 money_cols=[7],
                 date_cols=[8])

    return wb


def generate_export_bytes(tenant, days=None):
    """Generate Excel and return as bytes (for HTTP response or file save)."""
    wb = generate_export(tenant, days=days)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
